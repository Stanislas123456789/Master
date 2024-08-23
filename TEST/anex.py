import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Side

# Listes des termes financiers, des explications et des exemples détaillés
financial_terms = [
    "actors", "market participants", "limit order book", "bid-ask", "liquidity", "stocks", "dividends", "repo", "bonds", 
    "interest rates", "yield curve", "zero-coupon curve", "duration", "convexity", "term structure", "swap rates", "swaps", 
    "commodities", "convenience yield", "carry cost", "carry", "rolldown", "contango", "backwardation", "futures", 
    "margining", "non-arbitrage conditions", "replication", "credit", "spread", "options", "greeks", "risk", 
    "volatilities", "volatility surface"
]

explanations = [
    "Individuals or entities participating in financial markets.",
    "Participants who engage in the buying and selling of financial instruments.",
    "A record of all outstanding orders to buy and sell a particular financial instrument.",
    "The difference between the highest price a buyer is willing to pay and the lowest price a seller will accept.",
    "The ease with which an asset can be bought or sold in the market without affecting its price.",
    "Securities representing ownership in a corporation and a claim on part of its assets and earnings.",
    "Payments made by a corporation to its shareholders, usually as a distribution of profits.",
    "A form of short-term borrowing for dealers in government securities.",
    "Debt instruments issued by corporations or governments to raise capital.",
    "The cost of borrowing money, usually expressed as a percentage.",
    "A graph showing the relationship between bond yields and maturities.",
    "A curve that plots the yields of zero-coupon bonds against their maturities.",
    "A measure of the sensitivity of the price of a bond to a change in interest rates.",
    "A measure of the curvature of the relationship between bond prices and bond yields.",
    "The relationship between interest rates or bond yields and different terms or maturities.",
    "The interest rates applicable to swap contracts.",
    "Financial contracts in which two parties exchange streams of payments over time.",
    "Raw materials or primary agricultural products that can be bought and sold.",
    "The benefit or premium associated with holding a physical commodity over a futures contract.",
    "The costs associated with holding a financial position, including interest and storage costs.",
    "The net return on an investment after accounting for the costs of carrying the position.",
    "The change in the yield of a bond or futures contract as it approaches maturity.",
    "A situation where the futures price is higher than the expected future spot price.",
    "A situation where the futures price is lower than the expected future spot price.",
    "Financial contracts obligating the buyer to purchase an asset or the seller to sell an asset at a predetermined future date and price.",
    "The process of settling the profits and losses of futures contracts on a daily basis.",
    "Conditions where no arbitrage opportunities exist, ensuring fair pricing in the market.",
    "The process of mimicking the payoff of a financial instrument using a combination of other instruments.",
    "The risk of loss due to a borrower's failure to repay a loan or meet contractual obligations.",
    "The difference between the yield on a corporate bond and a government bond of similar maturity.",
    "Financial derivatives that provide the right, but not the obligation, to buy or sell an asset at a specified price.",
    "Measures of sensitivity of the price of options to changes in underlying parameters.",
    "The possibility of losing money on an investment or business venture.",
    "Statistical measures of the dispersion of returns for a given security or market index.",
    "A three-dimensional plot showing implied volatilities for different option strikes and expirations."
]

exemples_detailles = [
    "Speculators take on risk to profit from price changes. Hedgers use futures to protect against adverse price movements. Banks facilitate trades and provide liquidity.",
    "Traders, investors, brokers, and financial institutions participate in markets to buy and sell assets.",
    "An electronic list of buy and sell orders for a specific security, organized by price level.",
    "The bid is the highest price a buyer is willing to pay, while the ask is the lowest price a seller will accept.",
    "High liquidity means assets can be easily bought or sold without significantly affecting their prices.",
    "Shares of ownership in a company, representing a claim on part of its assets and earnings.",
    "Payments made by companies to their shareholders, typically from profits or reserves.",
    "Short-term borrowing where securities are used as collateral. Used by dealers in government securities.",
    "Issued by corporations or governments to raise funds. Investors earn interest until maturity.",
    "The cost of borrowing money, expressed as a percentage of the amount borrowed.",
    "A graphical representation showing the relationship between bond yields and maturities.",
    "Shows yields of zero-coupon bonds plotted against their respective maturities.",
    "Measures how sensitive a bond's price is to changes in interest rates.",
    "Measures the curvature of the relationship between bond prices and bond yields.",
    "Shows how interest rates or bond yields vary with different terms or maturities.",
    "Interest rates applicable to swap contracts, where parties exchange cash flows.",
    "Contracts where parties exchange financial streams based on asset prices over time.",
    "Primary goods that can be traded, like metals, oil, or agricultural products.",
    "Benefit of holding a physical commodity, such as convenience or usage benefits.",
    "Costs associated with maintaining a financial position, including interest and storage.",
    "Net return on an investment after accounting for costs associated with holding the position.",
    "Change in a bond's yield as it nears maturity, affecting its market price.",
    "Futures price is higher than the expected future spot price, indicating cost of carry.",
    "Futures price is lower than the expected future spot price, indicating income from carry.",
    "Contracts obliging parties to buy or sell assets at predetermined future dates and prices.",
    "Settlement of futures contracts' daily profits and losses, ensuring fair pricing.",
    "Conditions where no risk-free profit can be made from market discrepancies.",
    "Imitating payoff of one instrument with another to replicate its performance.",
    "Risk of loss due to borrower failing to meet repayment obligations.",
    "Difference between yields on corporate bonds and government bonds of similar maturity.",
    "Derivatives giving right to buy or sell assets at agreed prices, but not obligation.",
    "Sensitivities of option prices to changes in factors like underlying asset price and time.",
    "Risk of financial loss from investment or business activities.",
    "Shows option volatilities at different strikes and expiration dates, key for pricing and risk analysis."
]

# Créer un DataFrame avec les termes financiers, les explications et les exemples détaillés
df = pd.DataFrame({
    "Financial Terms": financial_terms,
    "Explanations": explanations,
    "Exemples détaillés": exemples_detailles
})

# Définir le chemin complet où le fichier Excel est enregistré
file_path = r"C:\Users\Smichel\Desktop\TEST\financial_terms_with_explanations.xlsx"

# Enregistrer le DataFrame dans un fichier Excel
df.to_excel(file_path, index=False)

# Charger le fichier Excel pour appliquer les styles de bordure
wb = load_workbook(filename=file_path)
ws = wb.active

# Définir un style de bordure mince
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Appliquer la bordure à toutes les cellules remplies dans les colonnes A, B et C
for row in ws.iter_rows(min_row=1, max_col=3, max_row=ws.max_row):
    for cell in row:
        if cell.value is not None:
            cell.border = thin_border

# Enregistrer le fichier Excel avec les bordures appliquées
wb.save(file_path)

print(f"Excel file with borders and detailed examples has been generated: {file_path}")
